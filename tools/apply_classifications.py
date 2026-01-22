#!/usr/bin/env python3
"""Apply classifications from CSV/XLSX back into app/data.json.

Usage:
  python3 apply_classifications.py --data app/data.json --in vehicle_classification_template.xlsx
"""
import argparse, json, shutil

def load_table(path):
    if path.lower().endswith(".csv"):
        import csv
        with open(path, newline="", encoding="utf-8") as f:
            return list(csv.DictReader(f))
    if path.lower().endswith(".xlsx"):
        from openpyxl import load_workbook
        wb = load_workbook(path)
        ws = wb["Classifications"]
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        out=[]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if all(v is None for v in row):
                continue
            out.append({headers[i]: ("" if row[i] is None else str(row[i])) for i in range(len(headers))})
        return out
    raise SystemExit("Input must be .csv or .xlsx")

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--data", default="app/data.json")
    ap.add_argument("--in", dest="inp", required=True)
    args = ap.parse_args()

    with open(args.data, "r", encoding="utf-8") as f:
        data = json.load(f)

    valid = set(data.get("vehicleClasses", []))
    questions = data.get("questions", [])
    by_id = {q.get("id",""): q for q in questions if q.get("id")}
    by_name = {q.get("answer",""): q for q in questions if q.get("answer")}

    table = load_table(args.inp)

    updated=0
    invalid=[]
    missing=[]

    for r in table:
        rid = (r.get("id") or "").strip()
        name = (r.get("name") or "").strip()
        cls = (r.get("class") or "").strip()

        if not cls:
            continue

        if cls not in valid:
            invalid.append((rid or name, cls))
            continue

        q = by_id.get(rid) if rid else by_name.get(name)
        if not q:
            missing.append(rid or name)
            continue

        if q.get("class") != cls:
            q["class"] = cls
            updated += 1

    bak = args.data + ".bak"
    shutil.copy2(args.data, bak)

    with open(args.data, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)

    print(f"Updated classifications: {updated}")
    if invalid:
        print("\nInvalid class values:")
        for k,v in invalid[:50]:
            print(f" - {k}: {v}")
    if missing:
        print("\nRows that did not match any question:")
        for k in missing[:50]:
            print(f" - {k}")
    print(f"Backup created: {bak}")

if __name__ == "__main__":
    main()
